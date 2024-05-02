import mysql.connector
from openpyxl import load_workbook


workbook = load_workbook(filename='Цены на комплектующие.xlsx')

workbook1 = load_workbook(filename='Вентлиторы.XLSX')

data_base = mysql.connector.connect(
    host='localhost',
    user="admin",
    password="Z1s2e3v4g5.",
    database="altair_materials"
)

cursor = data_base.cursor()

cursor.execute('''CREATE TABLE motors(
               id INT PRIMARY KEY AUTO_INCREMENT,
               name VARCHAR(60),
               manufacturer VARCHAR(60),
               article VARCHAR(30),
               parameter VARCHAR(30),
               price DECIMAL(10, 2),
               groups VARCHAR(30)
               )''')

cursor.execute('''CREATE TABLE contactors(
               id INT PRIMARY KEY AUTO_INCREMENT,
               name VARCHAR(60),
               manufacturer VARCHAR(60),
               article VARCHAR(30),
               parameter VARCHAR(30),
               price DECIMAL(10, 2),
               groups VARCHAR(20)
               )''')


cursor.execute('''CREATE TABLE FC(
               id INT PRIMARY KEY AUTO_INCREMENT,
               name VARCHAR(60),
               manufacturer VARCHAR(60),
               article VARCHAR(30),
               parameter VARCHAR(30),
               price DECIMAL(10, 2),
               groups VARCHAR(20)
               )''')

cursor.execute('''CREATE TABLE motor_protection(
               id INT PRIMARY KEY AUTO_INCREMENT,
               name VARCHAR(60),
               manufacturer VARCHAR(60),
               article VARCHAR(30),
               parameter VARCHAR(30),
               price DECIMAL(10, 2),
               groups VARCHAR(20)
               )''')

cursor.execute('''CREATE TABLE other(
               id INT PRIMARY KEY AUTO_INCREMENT,
               name VARCHAR(60),
               manufacturer VARCHAR(60),
               article VARCHAR(30),
               parameter VARCHAR(30),
               price DECIMAL(10, 2),
               groups VARCHAR(20)
               )''')

cursor.execute('''CREATE TABLE water_heater(
               id INT PRIMARY KEY AUTO_INCREMENT,
               name VARCHAR(60),
               manufacturer VARCHAR(60),
               article VARCHAR(30),
               parameter VARCHAR(30),
               price DECIMAL(10, 2),
               groups VARCHAR(20)
               )''')

cursor.execute('''CREATE TABLE electro_heater(
               id INT PRIMARY KEY AUTO_INCREMENT,
               name VARCHAR(60),
               manufacturer VARCHAR(60),
               article VARCHAR(30),
               parameter VARCHAR(30),
               price DECIMAL(10, 2),
               groups VARCHAR(20)
               )''')

cursor.execute('''CREATE TABLE plc(
               id INT PRIMARY KEY AUTO_INCREMENT,
               name VARCHAR(60),
               manufacturer VARCHAR(60),
               article VARCHAR(30),
               parameter VARCHAR(30),
               price DECIMAL(10, 2),
               groups VARCHAR(20)
               )''')

cursor.execute('''CREATE TABLE breakers(
               id INT PRIMARY KEY AUTO_INCREMENT,
               name VARCHAR(60),
               manufacturer VARCHAR(60),
               article VARCHAR(30),
               parameter VARCHAR(30),
               price DECIMAL(10, 2),
               groups VARCHAR(20)
               )''')

cursor.execute('''CREATE TABLE relay(
               id INT PRIMARY KEY AUTO_INCREMENT,
               name VARCHAR(60),
               manufacturer VARCHAR(60),
               article VARCHAR(30),
               parameter VARCHAR(30),
               price DECIMAL(10, 2),
               groups VARCHAR(20)
               )''')

cursor.execute('''CREATE TABLE cable(
               id INT PRIMARY KEY AUTO_INCREMENT,
               name VARCHAR(60),
               manufacturer VARCHAR(60),
               article VARCHAR(30),
               parameter VARCHAR(30),
               price DECIMAL(10, 2),
               groups VARCHAR(20)
               )''')

cursor.execute('''CREATE TABLE terminals(
               id INT PRIMARY KEY AUTO_INCREMENT,
               name VARCHAR(60),
               manufacturer VARCHAR(60),
               article VARCHAR(30),
               parameter VARCHAR(30),
               price DECIMAL(10, 2),
               groups VARCHAR(20)
               )''')

cursor.execute('''CREATE TABLE panels(
               id INT PRIMARY KEY AUTO_INCREMENT,
               name VARCHAR(60),
               manufacturer VARCHAR(60),
               article VARCHAR(30),
               parameter VARCHAR(30),
               price DECIMAL(10, 2),
               groups VARCHAR(20)
               )''')

print(workbook.get_sheet_names())
sheet_panels = workbook.get_sheet_by_name('ЩМП')
data_panels = []
for i in range(2, sheet_panels.max_column + 1):
    row = []
    for j in range(2, sheet_panels.max_row + 1):
        row.append(sheet_panels.cell(row=j, column=i).value)
    data_panels.append(row)
print(data_panels)


res_panels = [[0 for j in range(len(data_panels))] for i in range(len(data_panels[0]))]

for i in range(len(data_panels)):
    for j in range(len(data_panels[i])):
        res_panels[j][i] = data_panels[i][j]

print(res_panels)
for i in res_panels:
    sql = "INSERT INTO panels (name, manufacturer, article, parameter, price, groups) VALUES (%s, %s, %s, %s, %s, %s)"
    val = tuple(i)
    cursor.execute(sql, val)
data_base.commit()

sheet_contactors = workbook.get_sheet_by_name('Контакторы')
data_contactors = []
for i in range(2, sheet_contactors.max_column + 1):
    row = []
    for j in range(2, sheet_contactors.max_row + 1):
        row.append(sheet_contactors.cell(row=j, column=i).value)
    data_contactors.append(row)
print(data_contactors)


res_contactors = [[0 for j in range(len(data_contactors))] for i in range(len(data_contactors[0]))]

for i in range(len(data_contactors)):
    for j in range(len(data_contactors[i])):
        res_contactors[j][i] = data_contactors[i][j]

print(res_contactors)
for i in res_contactors:
    sql = "INSERT INTO contactors (name, manufacturer, article, parameter, price, groups) VALUES (%s, %s, %s, %s, %s, %s)"
    val = tuple(i)
    cursor.execute(sql, val)
data_base.commit()


sheet_plc = workbook.get_sheet_by_name('ПЛК')
data_plc = []
for i in range(2, sheet_plc.max_column + 1):
    row = []
    for j in range(2, sheet_plc. max_row+1 ):
        row.append(sheet_plc.cell(row=j, column=i).value)
    data_plc.append(row)
print(data_plc)

res_plc = [[0 for j in range(len(data_plc))] for i in range(len(data_plc[0]))]

for i in range(len(data_plc)):
    for j in range(len(data_plc[i])):
        res_plc[j][i] = data_plc[i][j]

print(res_plc)
for i in res_plc:
    sql = "INSERT INTO plc (name, manufacturer, article, parameter, price, groups) VALUES (%s, %s, %s, %s, %s, %s)"
    val = tuple(i)
    cursor.execute(sql, val)

data_base.commit()


sheet_breakers = workbook.get_sheet_by_name('Автоматические выключатели')
data_breakers = []
for i in range(2, sheet_breakers.max_column + 1):
    row = []
    for j in range(2, sheet_breakers. max_row+1 ):
        row.append(sheet_breakers.cell(row=j, column=i).value)
    data_breakers.append(row)
print(data_breakers)

res_breakers = [[0 for j in range(len(data_breakers))] for i in range(len(data_breakers[0]))]
print(res_breakers)
for i in range(len(data_breakers)):
    for j in range(len(data_breakers[i])):
        res_breakers[j][i] = data_breakers[i][j]

print(res_breakers)
for i in res_breakers:
    sql = "INSERT INTO breakers (name, manufacturer, article, parameter, price, groups) VALUES (%s, %s, %s, %s, %s, %s)"
    val = tuple(i)
    cursor.execute(sql, val)

data_base.commit()

sheet_relay = workbook.get_sheet_by_name('Реле')
data_relay = []
for i in range(2, sheet_relay.max_column + 1):
    row = []
    for j in range(2, sheet_relay. max_row+1 ):
        row.append(sheet_relay.cell(row=j, column=i).value)
    data_relay.append(row)
print(data_relay)

res_relay = [[0 for j in range(len(data_relay))] for i in range(len(data_relay[0]))]

for i in range(len(data_relay)):
    for j in range(len(data_relay[i])):
        res_relay[j][i] = data_relay[i][j]

print(res_relay)
for i in res_relay:
    sql = "INSERT INTO relay (name, manufacturer, article, parameter, price, groups) VALUES (%s, %s, %s, %s, %s, %s)"
    val = tuple(i)
    cursor.execute(sql, val)

data_base.commit()

sheet_cable = workbook.get_sheet_by_name('Кабель')
data_cable = []
for i in range(2, sheet_cable.max_column + 1):
    row = []
    for j in range(2, sheet_cable. max_row+1 ):
        row.append(sheet_cable.cell(row=j, column=i).value)
    data_cable.append(row)
print(data_cable)

res_cable = [[0 for j in range(len(data_cable))] for i in range(len(data_cable[0]))]

for i in range(len(data_cable)):
    for j in range(len(data_cable[i])):
        res_cable[j][i] = data_cable[i][j]

print(res_cable)
for i in res_cable:
    sql = "INSERT INTO cable (name, manufacturer, article, parameter, price, groups) VALUES (%s, %s, %s, %s, %s, %s)"
    val = tuple(i)
    cursor.execute(sql, val)

data_base.commit()

sheet_terminals = workbook.get_sheet_by_name('Клеммы')
data_terminals = []
for i in range(2, sheet_terminals.max_column + 1):
    row = []
    for j in range(2, sheet_terminals. max_row+1 ):
        row.append(sheet_terminals.cell(row=j, column=i).value)
    data_terminals.append(row)
print(data_terminals)


res_terminals = [[0 for j in range(len(data_terminals))] for i in range(len(data_terminals[0]))]

for i in range(len(data_terminals)):
    for j in range(len(data_terminals[i])):
        res_terminals[j][i] = data_terminals[i][j]

print(res_terminals)
for i in res_terminals:
    sql = "INSERT INTO terminals (name, manufacturer, article, parameter, price, groups) VALUES (%s, %s, %s, %s, %s, %s)"
    val = tuple(i)
    cursor.execute(sql, val)

data_base.commit()

print(workbook1.get_sheet_names())

sheet_motors = workbook1.get_sheet_by_name('Электродвигатели')
data_motors = []
for i in range(2, sheet_motors.max_column + 1):
    row = []
    for j in range(2, sheet_motors.max_row + 1):
        row.append(sheet_motors.cell(row=j, column=i).value)
    data_motors.append(row)
print(data_motors)


res_motors = [[0 for j in range(len(data_motors))] for i in range(len(data_motors[0]))]

for i in range(len(data_motors)):
    for j in range(len(data_motors[i])):
        res_motors[j][i] = data_motors[i][j]

print(res_motors)
for i in res_motors:
    sql = "INSERT INTO motors (name, manufacturer, article, parameter, price, groups) VALUES (%s, %s, %s, %s, %s, %s)"
    val = tuple(i)
    cursor.execute(sql, val)

data_base.commit()

sheet_motor_protection = workbook1.get_sheet_by_name('Автомат защ. дв.')
data_motor_protection = []
for i in range(2, sheet_motor_protection.max_column + 1):
    row = []
    for j in range(2, sheet_motor_protection.max_row + 1):
        row.append(sheet_motor_protection.cell(row=j, column=i).value)
    data_motor_protection.append(row)
print(data_motor_protection)


res_motor_protection = [[0 for j in range(len(data_motor_protection))] for i in range(len(data_motor_protection[0]))]

for i in range(len(data_motor_protection)):
    for j in range(len(data_motor_protection[i])):
        res_motor_protection[j][i] = data_motor_protection[i][j]

print(res_motor_protection)
for i in res_motor_protection:
    sql = "INSERT INTO motor_protection (name, manufacturer, article, parameter, price, groups) VALUES (%s, %s, %s, %s, %s, %s)"
    val = tuple(i)
    cursor.execute(sql, val)

data_base.commit()

sheet_FC = workbook1.get_sheet_by_name('ПЧ')
data_FC = []
for i in range(2, sheet_FC.max_column + 1):
    row = []
    for j in range(2, sheet_FC.max_row + 1):
        row.append(sheet_FC.cell(row=j, column=i).value)
    data_FC.append(row)
print(data_FC)


res_FC = [[0 for j in range(len(data_FC))] for i in range(len(data_FC[0]))]

for i in range(len(data_FC)):
    for j in range(len(data_FC[i])):
        res_FC[j][i] = data_FC[i][j]

print(res_FC)
for i in res_FC:
    sql = "INSERT INTO FC (name, manufacturer, article, parameter, price, groups) VALUES (%s, %s, %s, %s, %s, %s)"
    val = tuple(i)
    cursor.execute(sql, val)

data_base.commit()

sheet_heater_water = workbook1.get_sheet_by_name('Нагреватель водяной')
data_heater_water = []
for i in range(2, sheet_heater_water.max_column + 1):
    row = []
    for j in range(2, sheet_heater_water.max_row + 1):
        row.append(sheet_heater_water.cell(row=j, column=i).value)
    data_heater_water.append(row)
print(data_heater_water)


res_heater_water = [[0 for j in range(len(data_heater_water))] for i in range(len(data_heater_water[0]))]

for i in range(len(data_heater_water)):
    for j in range(len(data_heater_water[i])):
        res_heater_water[j][i] = data_heater_water[i][j]

print(res_heater_water)
for i in res_heater_water:
    sql = "INSERT INTO water_heater (name, manufacturer, article, parameter, price, groups) VALUES (%s, %s, %s, %s, %s, %s)"
    val = tuple(i)
    cursor.execute(sql, val)

data_base.commit()

sheet_heater_el = workbook1.get_sheet_by_name('Электро нагреватель')
data_heater_el = []
for i in range(2, sheet_heater_el.max_column + 1):
    row = []
    for j in range(2, sheet_heater_el.max_row + 1):
        row.append(sheet_heater_el.cell(row=j, column=i).value)
    data_heater_el.append(row)
print(data_heater_el)


res_heater_el = [[0 for j in range(len(data_heater_el))] for i in range(len(data_heater_el[0]))]

for i in range(len(data_heater_el)):
    for j in range(len(data_heater_el[i])):
        res_heater_el[j][i] = data_heater_el[i][j]

print(res_heater_el)
for i in res_heater_el:
    sql = "INSERT INTO electro_heater (name, manufacturer, article, parameter, price, groups) VALUES (%s, %s, %s, %s, %s, %s)"
    val = tuple(i)
    cursor.execute(sql, val)

data_base.commit()

sheet_other = workbook1.get_sheet_by_name('Другое')
data_other = []
for i in range(2, sheet_other.max_column + 1):
    row = []
    for j in range(2, sheet_other.max_row + 1):
        row.append(sheet_other.cell(row=j, column=i).value)
    data_other.append(row)
print(data_other)

res_other = [[0 for j in range(len(data_other))] for i in range(len(data_other[0]))]

for i in range(len(data_other)):
    for j in range(len(data_other[i])):
        res_other[j][i] = data_other[i][j]

print(res_other)
for i in res_other:
    sql = "INSERT INTO other (name, manufacturer, article, parameter, price, groups) VALUES (%s, %s, %s, %s, %s, %s)"
    val = tuple(i)
    cursor.execute(sql, val)

data_base.commit()

data_base.close()
