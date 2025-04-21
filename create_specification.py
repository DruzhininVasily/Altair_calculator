from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Side, Border, Alignment
from openpyxl.drawing.image import Image
from datetime import date
from fpdf import FPDF


def data_processing(data, static, rate):
    data_set = set(data)
    out_data = []
    # Подсчет количества элементов
    for j, i in enumerate(data_set):
        i = list(i)
        i.insert(-1, data.count(tuple(i)))
        i.insert(0, j + 1)
        i[-1] = round((float(i[-1]) * rate), 2)
        i.append(i[-2] * i[-1])
        out_data.append(i)
    # Добавление статических элементов
    index = out_data[-1][0]
    temp = [e.copy() for e in static]
    for i in range(len(temp)):
        index += 1
        temp[i].insert(0, index)
        temp[i][-1] = round((float(temp[i][-1]) * rate), 2)
        temp[i][-2] = round((float(temp[i][-2]) * rate), 2)
    out_data.extend(temp)
    # Разбитие элементов на группы
    panel_list = []
    kip_list = []
    for i in out_data:
        if 'Panel' in i:
            i.remove('Panel')
            panel_list.append(i)
        elif 'Kip' in i:
            i.remove('Kip')
            kip_list.append(i)
    return panel_list, kip_list


def create_ws(ws, title):
    exel_head = ['№', 'Наименование', 'Производитель', 'Артикул', 'Параметр', 'Код 1С', 'Количество', 'Цена шт.', 'Сумма']
    ws.title = title
    ws.append(exel_head)
    for item in range(1, ws.max_column + 1):
        ws.cell(row=1, column=item).fill = PatternFill(fgColor="006CCDEC", fill_type='solid')
    ws.column_dimensions['A'].width = 4.22
    ws.column_dimensions['B'].width = 33
    ws.column_dimensions['C'].width = 15.11
    ws.column_dimensions['D'].width = 24.78
    ws.column_dimensions['E'].width = 24.5
    ws.column_dimensions['F'].width = 12.90
    ws.column_dimensions['G'].width = 10.33
    ws.column_dimensions['H'].width = 10.33
    ws.column_dimensions['I'].width = 10.33
    return ws


def add_sum(ws, target_sum):
    cell = ws.cell(column=ws.max_column, row=ws.max_row + 3)
    cell.value = target_sum
    cell = ws.cell(column=ws.max_column - 1, row=ws.max_row)
    cell.value = 'Сумма: '


def create_tcp(sheet, sum_elements, kip_list, panel_price, shau_name):
    exel_head = ['№', 'Обозначение', 'Код 1С', 'Наименование', 'Кол-во', 'Ед.', 'Валюта', 'Цена', 'Сумма']
    image = Image('static/image/Tcp_head.png')
    image.height = 90
    image.width = 190
    sheet.add_image(image, 'B1')
    font_1 = Font(size=10, name='Tahoma')
    font_2 = Font(size=8, color='07B6F4', name='Tahoma')
    font_3 = Font(size=16, bold=True, name='Tahoma')
    font_4 = Font(size=16, name='Tahoma')
    font_5 = Font(size=10, bold=True, name='Tahoma')
    font_6 = Font(size=12, bold=True, name='Tahoma')
    font_7 = Font(size=8, bold=True, name='Tahoma')
    font_8 = Font(size=8, name='Tahoma')
    thins = Side(border_style="thin", color="000000")
    sheet.column_dimensions['A'].width = 1
    sheet.column_dimensions['B'].width = 3
    sheet.column_dimensions['C'].width = 15
    sheet.column_dimensions['D'].width = 15
    sheet.column_dimensions['E'].width = 22
    sheet.column_dimensions['F'].width = 9
    sheet.column_dimensions['G'].width = 8
    sheet.column_dimensions['H'].width = 8
    sheet.column_dimensions['I'].width = 8
    sheet.column_dimensions['J'].width = 8
    sheet.merge_cells('B5:C5')
    sheet.merge_cells('H7:J9')
    sheet.merge_cells('B7:C7')
    sheet.merge_cells('B8:C8')
    sheet.merge_cells('B15:J15')
    sheet.merge_cells('B18:J18')
    sheet.merge_cells('B19:J19')
    sheet.merge_cells('B22:J22')
    sheet.row_dimensions[5].height = 35
    sheet.row_dimensions[7].height = 23
    sheet.row_dimensions[8].height = 60
    sheet.row_dimensions[19].height = 42
    sheet['H7'].alignment = Alignment(vertical='top')
    sheet['B5'].alignment = Alignment(wrapText=True)
    sheet['B5'].value = '+7 495 641 16 01\ninfo@altair-gr.ru'
    sheet['B5'].font = font_1
    sheet['B6'].value = '123290, г. Москва, 1-й Магистральный тупик, дом 5 А'
    sheet['B6'].font = font_1
    sheet['B7'].hyperlink = 'https://www.altair-gr.ru'
    sheet['B7'].value = 'www.altair-gr.ru'
    sheet['B7'].font = font_2
    sheet['B8'].alignment = Alignment(wrapText=True)
    sheet['B8'].value = 'ООО «АЛТАИР ГРУПП»\nИНН 7714427104\nКПП 771401001'
    sheet['B8'].font = font_1
    sheet['B11'].value = 'Коммерческое предложение '
    sheet['B11'].font = font_3
    sheet['B12'].value = 'на поставку оборудования'
    sheet['B12'].font = font_4
    sheet['B13'].font = font_4
    sheet['B13'].value = '№ '
    sheet['B15'].value = 'Объект: '
    sheet['B15'].font = font_5
    sheet['B16'].value = 'Исполнитель: '
    sheet['B16'].font = font_5
    sheet['B18'].value = 'Уважаемый(ая) '
    sheet['B18'].font = font_6
    sheet['B19'].alignment = Alignment(wrapText=True)
    sheet['B19'].value = 'Благодарим за Ваш запрос.\nПредлагаем поставку оборудования ALTAIR согласно номенклатуре, ценам и коммерческим условиям, изложенным в настоящем предложении.'
    sheet['B19'].font = font_1
    sheet['B22'].value = 'Автоматика'
    sheet['B22'].font = font_7
    index = 1
    for i, j in enumerate(exel_head):
        sheet[21][i+1].value = j
        sheet[21][i+1].font = font_5
        sheet[21][i + 1].fill = PatternFill(fill_type='solid', start_color='D7F5F5', end_color='D7F5F5')
    for i, j in enumerate(kip_list):
        sheet[23 + i][1].value = index
        sheet[23 + i][1].font = font_8
        sheet[23 + i][2].value = j[3]
        sheet[23 + i][2].font = font_8
        sheet[23 + i][3].value = j[5]
        sheet[23 + i][3].font = font_8
        sheet[23 + i][4].alignment = Alignment(wrapText=True)
        sheet[23 + i][4].value = j[1]
        sheet[23 + i][4].font = font_8
        sheet[23 + i][5].value = j[6]
        sheet[23 + i][5].font = font_8
        sheet[23 + i][6].value = 'шт'
        sheet[23 + i][6].font = font_8
        sheet[23 + i][7].value = 'RUB'
        sheet[23 + i][7].font = font_8
        sheet[23 + i][8].value = j[7]
        sheet[23 + i][8].font = font_8
        sheet[23 + i][9].value = j[8]
        sheet[23 + i][9].font = font_8
        index += 1
    sheet[sheet.max_row + 1][1].value = index
    sheet[sheet.max_row][1].font = font_8
    sheet[sheet.max_row][2].alignment = Alignment(wrapText=True)
    sheet[sheet.max_row][2].value = shau_name
    sheet[sheet.max_row][2].font = font_8
    sheet[sheet.max_row][3].value = ''
    sheet[sheet.max_row][3].font = font_8
    sheet[sheet.max_row][4].alignment = Alignment(wrapText=True)
    sheet[sheet.max_row][4].value = 'Шкаф управления автоматикой с ПЛК'
    sheet[sheet.max_row][4].font = font_8
    sheet[sheet.max_row][5].value = 1
    sheet[sheet.max_row][5].font = font_8
    sheet[sheet.max_row][6].value = 'шт'
    sheet[sheet.max_row][6].font = font_8
    sheet[sheet.max_row][7].value = 'RUB'
    sheet[sheet.max_row][7].font = font_8
    sheet[sheet.max_row][8].value = panel_price
    sheet[sheet.max_row][8].font = font_8
    sheet[sheet.max_row][9].value = panel_price
    sheet[sheet.max_row][9].font = font_8
    for row in range(21, sheet.max_row + 1):
        for col in range(1, sheet.max_column):
            sheet[row][col].border = Border(top=thins, bottom=thins, left=thins, right=thins)
    sheet[sheet.max_row + 2][9].value = f'Итого: {sum_elements:.2f}'
    sheet[sheet.max_row][9].font = font_5
    sheet[sheet.max_row][9].alignment = Alignment(horizontal='right')
    sheet[sheet.max_row + 1][9].value = f'В том числе НДС 20% {sum_elements * 0.2:.2f}'
    sheet[sheet.max_row][9].font = font_5
    sheet[sheet.max_row][9].alignment = Alignment(horizontal='right')
    sheet[sheet.max_row + 5][2].value = 'Коммерческие условия: '
    sheet[sheet.max_row][2].font = font_5
    sheet[sheet.max_row + 2][2].value = '1. Срок поставки: '
    sheet[sheet.max_row][2].font = font_1
    sheet[sheet.max_row][6].value = 'По договору'
    sheet[sheet.max_row][6].font = font_1
    sheet[sheet.max_row + 1][2].value = '2. Условия поставки: '
    sheet[sheet.max_row][2].font = font_1
    sheet[sheet.max_row][6].value = 'Отгрузка со склада АГ в Липецке'
    sheet[sheet.max_row][6].font = font_1
    sheet[sheet.max_row + 1][2].value = '3. Условия оплаты: '
    sheet[sheet.max_row][2].font = font_1
    sheet[sheet.max_row][6].value = 'По договору'
    sheet[sheet.max_row][6].font = font_1
    sheet[sheet.max_row + 1][2].value = '4. Механическая гарантия: '
    sheet[sheet.max_row][2].font = font_1
    sheet[sheet.max_row][6].value = '24 месяца с момента отгрузки.'
    sheet[sheet.max_row][6].font = font_1
    sheet[sheet.max_row + 1][2].value = '5. Предложение действительно до: '
    sheet[sheet.max_row][2].font = font_1
    sheet[sheet.max_row][6].value = ''
    sheet[sheet.max_row][6].font = font_1
    sheet[sheet.max_row + 2][2].value = 'С уважением, '
    sheet[sheet.max_row][2].font = font_1


def add_tcp_name(name, comp_object, number, customer, executor, date_now):
    workbook = load_workbook(filename='TCP.xlsx')
    sheet = workbook.get_sheet_by_name('Sheet')
    sheet['H7'].value = comp_object
    sheet['B13'].value += number
    sheet['B13'].value += f' от {date_now}'
    sheet['B15'].value += name
    sheet['B16'].value += executor
    sheet['B18'].value += customer
    sheet[sheet.max_row + 2][2].value = executor
    sheet[sheet.max_row + 1][2].value = 'Тел.: +7 495 6411601'
    workbook.save('TCP.xlsx')


def create_exel(data, static, rate, shau_name):
    wb = Workbook()
    panel_list, kip_list = data_processing(data, static, rate)
    # Запись элементов на в файл
    ws = wb.active
    create_ws(ws, 'Шкаф')
    sum_panel = 0
    kip_sum = 0
    for row in panel_list:
        sum_panel += float(row[-1])
        if row[-3] != 0:
            ws.append(row)
    add_sum(ws, sum_panel)
    if len(kip_list) > 0:
        ws1 = wb.create_sheet()
        create_ws(ws1, 'КИПиА')
        for row in kip_list:
            kip_sum += float(row[-1])
            ws1.append(row)
        add_sum(ws1, kip_sum)
    wb.save('specification.xlsx')
    wb1 = Workbook()
    sheet = wb1.active
    create_tcp(sheet, (sum_panel+kip_sum), kip_list, sum_panel, shau_name)
    wb1.save('TCP.xlsx')


def create_pdf(data):
    short_img = [
        'fire', 'damp_in', 'damp_out', 'switch_type', 'termostat',
        'humid', 'hood', 'signal_work', 'signal_alarm', 'remote_control', 'cooler'
    ]
    long_img = [
        'water', 'el_heater', 'in_out', 'in_in'
    ]
    pdf = FPDF(orientation='landscape')
    pdf.add_page()
    x = 2
    for i in data:
        short = False
        long = False
        for j in short_img:
            if j in i:
                short = True
        else:
            for n in long_img:
                if n in i:
                    long = True
        if short:
            w = 7
            step = 7
        elif long:
            w = 17
            step = 17
        else:
            w = 12
            step = 12
        if i != '':
            pdf.image(i, x=x, y=60, w=w, h=70)
            x += step
    pdf.output('schema.pdf')


def get_scheme(sys_type, param, num_param):
    file_name = ''
    quant = 1
    if sys_type == 'In_out':
        file_name += 'In_'
        if 'vent_quant_in' in num_param or 'vent_quant_out' in num_param:
            if int(num_param['vent_quant_in']) != 0 or int(num_param['vent_quant_out']) != 0:
                quant = 2
        if quant == 2:
            file_name += '2_motor_Heater_Water.pdf'
        else:
            file_name += '1_motor_'
            if 'first_heater_type' in param:
                file_name += 'Heater_Electro_2_steps.pdf'
            else:
                file_name += 'Heater_Water'
                if 'select_1' in param:
                    file_name += '_Cooler_Freon'
                    if 'filt_in' in param or 'filt_out' in param:
                        file_name += '_2_filters.pdf'
                    else:
                        file_name += '.pdf'
                else:
                    file_name += '.pdf'
    elif sys_type == 'In':
        file_name += 'In_'
        if 'vent_quant_in' in num_param:
            if int(num_param['vent_quant_in']) != 0:
                quant = 2
        if quant == 2:
            file_name += '2_motor_Heater_Water.pdf'
        else:
            file_name += '1_motor_'
            if 'first_heater_type' in param:
                file_name += 'Heater_Electro_2_steps.pdf'
            else:
                file_name += 'Heater_Water'
                if 'select_1' in param:
                    file_name += '_Cooler_Freon'
                    if 'filt_in' in param:
                        file_name += '_2_filters.pdf'
                    else:
                        file_name += '.pdf'
                else:
                    file_name += '.pdf'
    else:
        file_name += 'Out_'
        if 'vent_quant_out' in num_param:
            if int(num_param['vent_quant_out']) != 0:
                quant = 2
        if quant == 2:
            file_name += '2_motor_FC.pdf'
        else:
            file_name += '1_motor_FC.pdf'
    return file_name


def get_shau_name(obj):
    power = {0: "0.75", 1: "1.5", 2: "2.2", 3: "4", 4: "5.5", 5: "7.5", 6: "11", 7: "15", 8: "18.5", 9: "22", 10: "30", 11: "37", 12: "45", 13: '55'}
    heat_power = {0: '3', 1: '6', 2: '9', 3: '12', 4: '15', 5: '18', 6: '24'}
    name = 'AMC-'
    print(obj.numer_param)
    print(obj.parameters)
    if obj.system_type == 'In_out':
        name += 'AER.'
        quantity_1 = 1
        quantity_2 = 1
        if 'vent_quant_in' in obj.numer_param:
            quantity_1 += int(obj.numer_param['vent_quant_in'])
        if 'vent_quant_out' in obj.numer_param:
            quantity_2 += int(obj.numer_param['vent_quant_out'])
        name += str(quantity_1) + '.'
        if 'voltage_vent_in' not in obj.parameters:
            name += 'D'
        else:
            name += 'E'
        vent_power_1 = ''
        vent_power_2 = ''
        if 'vent_power_in' in obj.numer_param:
            vent_power_1 = power[int(obj.numer_param['vent_power_in'])]
        if 'vent_power_out' in obj.numer_param:
            vent_power_2 = power[int(obj.numer_param['vent_power_out'])]
        vent_power_1 = str(int(float(vent_power_1) * 100))
        if len(vent_power_1) < 4:
            vent_power_1 = '0' * (4-len(vent_power_1)) + vent_power_1
        vent_power_2 = str(int(float(vent_power_2) * 100))
        if len(vent_power_2) < 4:
            vent_power_2 = '0' * (4-len(vent_power_2)) + vent_power_2
        name += vent_power_1 + '.'
        name += str(quantity_2) + '.'
        if 'voltage_vent_out' not in obj.parameters:
            name += 'D'
        else:
            name += 'E'
        name += vent_power_2
    elif obj.system_type == 'In':
        name += 'AER.'
        quantity = 1
        if 'vent_quant_in' in obj.numer_param:
            quantity += int(obj.numer_param['vent_quant_in'])
        name += str(quantity) + '.'
        if 'voltage_vent_in' not in obj.parameters:
            name += 'D'
        else:
            name += 'E'
        vent_power_1 = ''
        if 'vent_power_in' in obj.numer_param:
            vent_power_1 = power[int(obj.numer_param['vent_power_in'])]
        vent_power_1 = str(int(float(vent_power_1) * 100))
        if len(vent_power_1) < 4:
            vent_power_1 = '0' * (4-len(vent_power_1)) + vent_power_1
        name += vent_power_1
    elif obj.system_type == 'Out':
        name += 'EXH.'
        quantity_2 = 1
        if 'vent_quant_out' in obj.numer_param:
            quantity_2 += int(obj.numer_param['vent_quant_out'])
        name += str(quantity_2)
        if 'voltage_vent_out' not in obj.parameters:
            name += 'D'
        else:
            name += 'E'
        vent_power_2 = ''
        if 'vent_power_out' in obj.numer_param:
            vent_power_2 = power[int(obj.numer_param['vent_power_out'])]
        vent_power_2 = str(int(float(vent_power_2) * 100))
        if len(vent_power_2) < 4:
            vent_power_2 = '0' * (4-len(vent_power_2)) + vent_power_2
        name += vent_power_2
    steps_1 = ''
    steps_2 = ''
    el_heat_1 = ''
    el_heat_2 = ''
    if 'first_heater_type' in obj.parameters and 'second_heat_type' in obj.parameters:
        name += '.EL.'
        steps_1 = int(obj.numer_param['first_steps']) + 1
        steps_2 = int(obj.numer_param['second_steps']) + 1
        el_heat_1 = heat_power[int(obj.numer_param['electrical_first_step_power'])]
        el_heat_2 = heat_power[int(obj.numer_param['electrical_second_step_power'])]
        el_heat_1 = str(int(el_heat_1) * 100)
        el_heat_2 = str(int(el_heat_2) * 100)
        if len(el_heat_1) < 4:
            el_heat_1 = '.' + '0' * (4-len(el_heat_1)) + el_heat_1
        elif len(el_heat_1) == 4:
            el_heat_1 += '.'
        if len(el_heat_2) < 4:
            el_heat_2 = '.' + '0' * (4-len(el_heat_2)) + el_heat_2
        elif len(el_heat_2) == 4:
            el_heat_2 += '.'
        name += str(steps_1)
        name += el_heat_1 + '.'
        name += str(steps_2)
        name += el_heat_2
    elif 'first_heater_type' in obj.parameters:
        name += '.EL.'
        steps_1 = int(obj.numer_param['first_steps']) + 1
        el_heat_1 = heat_power[int(obj.numer_param['electrical_first_step_power'])]
        el_heat_1 = str(int(el_heat_1) * 100)
        if len(el_heat_1) < 4:
            el_heat_1 = '.' + '0' * (4 - len(el_heat_1)) + el_heat_1
        elif len(el_heat_1) == 4:
            el_heat_1 += '.'
        name += str(steps_1)
        name += el_heat_1
    elif 'second_heat_type' in obj.parameters:
        name += '.EL.'
        steps_2 = int(obj.numer_param['second_steps']) + 1
        el_heat_2 = heat_power[int(obj.numer_param['electrical_second_step_power'])]
        el_heat_2 = str(int(el_heat_2) * 100)
        if len(el_heat_2) < 4:
            el_heat_2 = '.' + '0' * (4 - len(el_heat_2)) + el_heat_2
        elif len(el_heat_2) == 4:
            el_heat_2 += '.'
        name += str(steps_2)
        name += el_heat_2
    print(name)
    # Если есть ККБ, увлажнитель, осушение, УФ секции, в конце ставим V2 (не стандартное исполнение)
    return name


def create_general_tcp(systems, rate, system_names, tcp_data, quantitys):
    wb = Workbook()
    sheet = wb.active
    sheet.title = 'ТКП'
    exel_head = ['№', 'Обозначение', 'Код 1С', 'Наименование', 'Кол-во', 'Ед.', 'Валюта', 'Цена', 'Сумма']
    image = Image('static/image/Tcp_head.png')
    image.height = 90
    image.width = 190
    sheet.add_image(image, 'B1')
    font_1 = Font(size=10, name='Tahoma')
    font_2 = Font(size=8, color='07B6F4', name='Tahoma')
    font_3 = Font(size=16, bold=True, name='Tahoma')
    font_4 = Font(size=16, name='Tahoma')
    font_5 = Font(size=10, bold=True, name='Tahoma')
    font_6 = Font(size=12, bold=True, name='Tahoma')
    font_7 = Font(size=8, bold=True, name='Tahoma')
    font_8 = Font(size=8, name='Tahoma')
    thins = Side(border_style="thin", color="000000")
    sheet.column_dimensions['A'].width = 1
    sheet.column_dimensions['B'].width = 3
    sheet.column_dimensions['C'].width = 15
    sheet.column_dimensions['D'].width = 15
    sheet.column_dimensions['E'].width = 22
    sheet.column_dimensions['F'].width = 9
    sheet.column_dimensions['G'].width = 8
    sheet.column_dimensions['H'].width = 8
    sheet.column_dimensions['I'].width = 8
    sheet.column_dimensions['J'].width = 8
    sheet.merge_cells('B5:C5')
    sheet.merge_cells('H7:J9')
    sheet.merge_cells('B7:C7')
    sheet.merge_cells('B8:C8')
    sheet.merge_cells('B15:J15')
    sheet.merge_cells('B18:J18')
    sheet.merge_cells('B19:J19')
    sheet.row_dimensions[5].height = 35
    sheet.row_dimensions[7].height = 23
    sheet.row_dimensions[8].height = 60
    sheet.row_dimensions[19].height = 42
    sheet['H7'].alignment = Alignment(vertical='top')
    sheet['B5'].alignment = Alignment(wrapText=True)
    sheet['B5'].value = '+7 495 641 16 01\ninfo@altair-gr.ru'
    sheet['B5'].font = font_1
    sheet['B6'].value = '123290, г. Москва, 1-й Магистральный тупик, дом 5 А'
    sheet['B6'].font = font_1
    sheet['B7'].hyperlink = 'https://www.altair-gr.ru'
    sheet['B7'].value = 'www.altair-gr.ru'
    sheet['B7'].font = font_2
    sheet['B8'].alignment = Alignment(wrapText=True)
    sheet['B8'].value = 'ООО «АЛТАИР ГРУПП»\nИНН 7714427104\nКПП 771401001'
    sheet['B8'].font = font_1
    sheet['B11'].value = 'Коммерческое предложение '
    sheet['B11'].font = font_3
    sheet['B12'].value = 'на поставку оборудования'
    sheet['B12'].font = font_4
    sheet['B13'].font = font_4
    sheet['B13'].value = '№ '
    sheet['B15'].value = 'Объект: '
    sheet['B15'].font = font_5
    sheet['B16'].value = 'Исполнитель: '
    sheet['B16'].font = font_5
    sheet['B18'].value = 'Уважаемый(ая) '
    sheet['B18'].font = font_6
    sheet['B19'].alignment = Alignment(wrapText=True)
    sheet[
        'B19'].value = 'Благодарим за Ваш запрос.\nПредлагаем поставку оборудования ALTAIR согласно номенклатуре, ценам и коммерческим условиям, изложенным в настоящем предложении.'
    sheet['B19'].font = font_1
    sheet['H7'].value = tcp_data['company_object']
    sheet['B13'].value += str(tcp_data['tcp_number'])
    sheet['B13'].value += f' от {date.today()}'
    sheet['B15'].value += tcp_data['company_name']
    sheet['B16'].value += tcp_data['executor']
    sheet['B18'].value += tcp_data['customer']
    index = 1
    sum_elements = 0
    for i, j in enumerate(exel_head):
        sheet[21][i + 1].value = j
        sheet[21][i + 1].font = font_5
        sheet[21][i + 1].fill = PatternFill(fill_type='solid', start_color='D7F5F5', end_color='D7F5F5')
    for n, system in enumerate(systems):
        for q in range(int(quantitys[n])):
            sum_panel = 0
            system['spec'] = list(map(lambda x: tuple(x), system['spec']))
            panel_list, kip_list = data_processing(system['spec'], system['static'], rate)
            for row in panel_list:
                sum_panel += float(row[-1])
            sum_elements += sum_panel
            sheet.merge_cells(start_row=sheet.max_row+1, start_column=2, end_row=sheet.max_row+1, end_column=10)
            sheet[sheet.max_row][1].value = system_names[n]
            sheet[sheet.max_row][1].font = font_7
            for j in kip_list:
                sheet[sheet.max_row+1][1].value = index
                sheet[sheet.max_row][1].font = font_8
                sheet[sheet.max_row][2].value = j[3]
                sheet[sheet.max_row][2].font = font_8
                sheet[sheet.max_row][3].value = j[5]
                sheet[sheet.max_row][3].font = font_8
                sheet[sheet.max_row][4].alignment = Alignment(wrapText=True)
                sheet[sheet.max_row][4].value = j[1]
                sheet[sheet.max_row][4].font = font_8
                sheet[sheet.max_row][5].value = j[6]
                sheet[sheet.max_row][5].font = font_8
                sheet[sheet.max_row][6].value = 'шт'
                sheet[sheet.max_row][6].font = font_8
                sheet[sheet.max_row][7].value = 'RUB'
                sheet[sheet.max_row][7].font = font_8
                sheet[sheet.max_row][8].value = j[7]
                sheet[sheet.max_row][8].font = font_8
                sheet[sheet.max_row][9].value = j[8]
                sheet[sheet.max_row][9].font = font_8
                sum_elements += j[8]
                index += 1
            sheet[sheet.max_row + 1][1].value = index
            sheet[sheet.max_row][1].font = font_8
            sheet[sheet.max_row][2].alignment = Alignment(wrapText=True)
            sheet[sheet.max_row][2].value = system['shau_name']
            sheet[sheet.max_row][2].font = font_8
            sheet[sheet.max_row][3].value = ''
            sheet[sheet.max_row][3].font = font_8
            sheet[sheet.max_row][4].alignment = Alignment(wrapText=True)
            sheet[sheet.max_row][4].value = 'Шкаф управления автоматикой с ПЛК'
            sheet[sheet.max_row][4].font = font_8
            sheet[sheet.max_row][5].value = 1
            sheet[sheet.max_row][5].font = font_8
            sheet[sheet.max_row][6].value = 'шт'
            sheet[sheet.max_row][6].font = font_8
            sheet[sheet.max_row][7].value = 'RUB'
            sheet[sheet.max_row][7].font = font_8
            sheet[sheet.max_row][8].value = sum_panel
            sheet[sheet.max_row][8].font = font_8
            sheet[sheet.max_row][9].value = sum_panel
            sheet[sheet.max_row][9].font = font_8
            index += 1
    for row in range(21, sheet.max_row + 1):
        for col in range(1, sheet.max_column):
            sheet[row][col].border = Border(top=thins, bottom=thins, left=thins, right=thins)
    sheet[sheet.max_row + 2][9].value = f'Итого: {sum_elements:.2f}'
    sheet[sheet.max_row][9].font = font_5
    sheet[sheet.max_row][9].alignment = Alignment(horizontal='right')
    sheet[sheet.max_row + 1][9].value = f'В том числе НДС 20% {sum_elements * 0.2:.2f}'
    sheet[sheet.max_row][9].font = font_5
    sheet[sheet.max_row][9].alignment = Alignment(horizontal='right')
    sheet[sheet.max_row + 5][2].value = 'Коммерческие условия: '
    sheet[sheet.max_row][2].font = font_5
    sheet[sheet.max_row + 2][2].value = '1. Срок поставки: '
    sheet[sheet.max_row][2].font = font_1
    sheet[sheet.max_row][6].value = 'По договору'
    sheet[sheet.max_row][6].font = font_1
    sheet[sheet.max_row + 1][2].value = '2. Условия поставки: '
    sheet[sheet.max_row][2].font = font_1
    sheet[sheet.max_row][6].value = 'Отгрузка со склада АГ в Липецке'
    sheet[sheet.max_row][6].font = font_1
    sheet[sheet.max_row + 1][2].value = '3. Условия оплаты: '
    sheet[sheet.max_row][2].font = font_1
    sheet[sheet.max_row][6].value = 'По договору'
    sheet[sheet.max_row][6].font = font_1
    sheet[sheet.max_row + 1][2].value = '4. Механическая гарантия: '
    sheet[sheet.max_row][2].font = font_1
    sheet[sheet.max_row][6].value = '24 месяца с момента отгрузки.'
    sheet[sheet.max_row][6].font = font_1
    sheet[sheet.max_row + 1][2].value = '5. Предложение действительно до: '
    sheet[sheet.max_row][2].font = font_1
    sheet[sheet.max_row][6].value = ''
    sheet[sheet.max_row][6].font = font_1
    sheet[sheet.max_row + 2][2].value = 'С уважением, '
    sheet[sheet.max_row][2].font = font_1
    sheet[sheet.max_row + 2][2].value = tcp_data['executor']
    sheet[sheet.max_row + 1][2].value = 'Тел.: +7 495 6411601'
    wb.save('general_tcp.xlsx')

